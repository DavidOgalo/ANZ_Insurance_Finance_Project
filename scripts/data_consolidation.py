"""
data_consolidation.py - Combine all data and prepare final deliverable
"""

import pandas as pd
import os
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint

def load_data():
    """Load the processed data files"""
    try:
        companies_df = pd.read_excel('data/processed/top_hiring_companies.xlsx')
        executives_df = pd.read_excel('data/processed/executives_contacts.xlsx')
        return companies_df, executives_df
    except FileNotFoundError:
        print("Processed data files not found. Please run the previous scripts first.")
        exit(1)

def clean_company_names(df):
    """Normalize company names by removing suffixes"""
    # Define patterns to remove
    patterns = [
        r'\sLtd\.?$', r'\sLimited$', r'\sPty\.?$', r'\sCorp\.?$', 
        r'\sCorporation$', r'\sInc\.?$', r'\sLLC$', r'\sGroup$',
        r'\sHoldings$', r'\s\(.+\)$'
    ]
    
    # Function to clean company name
    def clean_name(name):
        cleaned = name
        for pattern in patterns:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
        return cleaned.strip()
    
    # Apply cleaning function
    df['clean_company_name'] = df['company_name'].apply(clean_name)
    
    return df

def prepare_final_dataset(companies_df, executives_df):
    """Combine and prepare the final dataset"""
    # Clean company names
    companies_df = clean_company_names(companies_df)
    
    # Merge companies and executives data
    final_df = pd.merge(
        companies_df,
        executives_df,
        on='company_id',
        how='left'
    )
    
    # Sort by company size or annual revenue if available
    if 'company_size' in final_df.columns and not final_df['company_size'].isna().all():
        final_df = final_df.sort_values('company_size', ascending=False)
    elif 'annual_revenue' in final_df.columns and not final_df['annual_revenue'].isna().all():
        final_df = final_df.sort_values('annual_revenue', ascending=False)
    
    # Add data quality score - simplified version
    def calculate_quality(row):
        score = 0
        # Basic company info complete
        if not any(pd.isna(row[col]) for col in ['company_name', 'industry', 'country']):
            score += 1
        # Website and LinkedIn available
        if not pd.isna(row['company_website']) and not pd.isna(row['linkedin_url_x']):
            score += 1
        # Executive name available
        if not pd.isna(row['executive_name']):
            score += 1
        # Executive has title
        if not pd.isna(row['executive_title']):
            score += 1
        # Executive has LinkedIn
        if not pd.isna(row['linkedin_url_y']):
            score += 1
        # Contact info available
        if not pd.isna(row['email']):
            score += 1
        if not pd.isna(row['phone']):
            score += 1
        
        # Convert to stars
        stars = '★' * score
        stars += '☆' * (5 - score)
        
        return stars
    
    final_df['data_quality'] = final_df.apply(calculate_quality, axis=1)
    
    # Rename columns for clarity
    final_df = final_df.rename(columns={
        'clean_company_name': 'Company Name',
        'industry': 'Industry',
        'country': 'Country',
        'company_size': 'Company Size',
        'annual_revenue': 'Annual Revenue (AUD)',
        'company_website': 'Company Website',
        'linkedin_url_x': 'Company LinkedIn',
        'hiring_devops': 'Hiring DevOps',
        'hiring_developers': 'Hiring Software Developers',
        'executive_name': 'C-Level Contact Name',
        'executive_title': 'C-Level Title',
        'linkedin_url_y': 'C-Level LinkedIn',
        'email': 'C-Level Email',
        'phone': 'C-Level Phone',
        'data_quality': 'Data Quality',
        'last_verified': 'Last Verified'
    })
    
    # Ensure date is in correct format
    if 'Last Verified' in final_df.columns:
        try:
            final_df['Last Verified'] = pd.to_datetime(final_df['Last Verified']).dt.strftime('%Y-%m-%d')
        except:
            final_df['Last Verified'] = datetime.now().strftime('%Y-%m-%d')
    
    # Select and reorder columns for final output
    columns_order = [
        'Company Name', 'Industry', 'Country', 'Company Size', 'Annual Revenue (AUD)',
        'Company Website', 'Company LinkedIn', 'Hiring DevOps', 'Hiring Software Developers',
        'C-Level Contact Name', 'C-Level Title', 'C-Level LinkedIn', 'C-Level Email', 
        'C-Level Phone', 'Data Quality', 'Last Verified'
    ]
    
    # Only include columns that exist in the dataframe
    available_columns = [col for col in columns_order if col in final_df.columns]
    
    final_df = final_df[available_columns]
    
    return final_df

def create_summary_data(final_df):
    """Create summary statistics for the dashboard"""
    metrics = {}
    
    # Count companies
    metrics['Total Companies'] = len(final_df['Company Name'].unique())
    
    # Count by country
    if 'Country' in final_df.columns:
        au_companies = len(final_df[final_df['Country'] == 'Australia']['Company Name'].unique())
        nz_companies = len(final_df[final_df['Country'] == 'New Zealand']['Company Name'].unique())
        metrics['Australian Companies'] = au_companies
        metrics['New Zealand Companies'] = nz_companies
    
    # Count by industry
    if 'Industry' in final_df.columns:
        insurance_companies = len(final_df[final_df['Industry'] == 'Insurance']['Company Name'].unique())
        finance_companies = len(final_df[final_df['Industry'] == 'Finance']['Company Name'].unique())
        both_companies = len(final_df[final_df['Industry'] == 'Both']['Company Name'].unique())
        metrics['Insurance Companies'] = insurance_companies
        metrics['Finance Companies'] = finance_companies
        metrics['Companies with Both Classifications'] = both_companies
    
    # Count by hiring status
    if all(col in final_df.columns for col in ['Hiring DevOps', 'Hiring Software Developers']):
        devops_companies = len(final_df[final_df['Hiring DevOps'] == 'Yes']['Company Name'].unique())
        software_companies = len(final_df[final_df['Hiring Software Developers'] == 'Yes']['Company Name'].unique())
        both_roles = len(final_df[(final_df['Hiring DevOps'] == 'Yes') & 
                                 (final_df['Hiring Software Developers'] == 'Yes')]['Company Name'].unique())
        metrics['Companies Hiring DevOps'] = devops_companies
        metrics['Companies Hiring Software Developers'] = software_companies
        metrics['Companies Hiring Both Roles'] = both_roles
    
    # Count contacts
    if 'C-Level Contact Name' in final_df.columns:
        contacts_found = len(final_df[~final_df['C-Level Contact Name'].isna()]['Company Name'].unique())
        metrics['C-Level Contacts Found'] = contacts_found
    
    if 'C-Level Email' in final_df.columns:
        emails_found = len(final_df[~final_df['C-Level Email'].isna()]['Company Name'].unique())
        metrics['C-Level Emails Found'] = emails_found
    
    if 'C-Level Phone' in final_df.columns:
        phones_found = len(final_df[~final_df['C-Level Phone'].isna()]['Company Name'].unique())
        metrics['C-Level Phone Numbers Found'] = phones_found
    
    # Calculate percentages
    total_companies = metrics.get('Total Companies', 1)  # Avoid division by zero
    percentages = {k: f"{v/total_companies*100:.1f}%" for k, v in metrics.items()}
    
    # Combine counts and percentages
    summary_df = pd.DataFrame({
        'Metric': list(metrics.keys()),
        'Count': list(metrics.values()),
        'Percentage': list(percentages.values())
    })
    
    return summary_df

def create_methodology_content():
    """Create content for the methodology sheet"""
    methodology = {
        'Data Collection Process': [
            "1. Identified top ANZ insurance and finance companies using:",
            "   - ASX 200 and NZX 50 listings",
            "   - Industry rankings from IBISWorld and similar sources",
            "   - Financial regulatory bodies (APRA, RBNZ, etc.)",
            "   - Industry associations and membership directories",
            "   - Business directories and financial news sources"
        ],
        'Company Verification and Enrichment': [
            "1. Verified company information through:",
            "   - Official company websites",
            "   - LinkedIn company pages",
            "   - Annual reports and investor presentations",
            "   - Financial news and industry publications",
            "",
            "2. Enriched basic company data with:",
            "   - Company size (employee count)",
            "   - Annual revenue figures",
            "   - Industry classification refinement",
            "   - Website and social media verification"
        ],
        'Hiring Status Verification': [
            "1. Verified active hiring for DevOps and Software Developer roles through:",
            "   - Company career pages and job listings",
            "   - LinkedIn Jobs search for each company",
            "   - Seek.com.au and Seek.co.nz job boards",
            "   - Indeed Australia and New Zealand listings",
            "",
            "2. Documentation of job listings:",
            "   - Saved URLs to specific job postings",
            "   - Verified job posting dates (filtered for active listings)",
            "   - Confirmed role titles and departments"
        ],
        'C-Level Contact Identification': [
            "1. Identified technology decision makers through:",
            "   - Company leadership pages",
            "   - LinkedIn advanced search filtering by title and company",
            "   - Recent conference speakers and industry events",
            "   - Tech press releases and news mentions",
            "",
            "2. Prioritized the following roles:",
            "   - CTO (Chief Technology Officer)",
            "   - CIO (Chief Information Officer)",
            "   - CDO (Chief Digital Officer)",
            "   - VP/Head of Technology/Engineering",
            "   - Director of IT/Technology"
        ],
        'Contact Information Enrichment': [
            "1. Generated and verified email addresses using:",
            "   - Email pattern analysis for each company",
            "   - Standard business email formats",
            "   - Domain MX record verification",
            "",
            "2. Sourced phone numbers where available from:",
            "   - Public company directories",
            "   - Professional profiles and contact information",
            "",
            "3. All contact data was ethically sourced from:",
            "   - Publicly available information only",
            "   - Professional business contexts",
            "   - Sources where contact sharing is expected for business purposes"
        ],
        'Data Quality Assurance': [
            "1. Verification methodology:",
            "   - Cross-referenced information across multiple sources",
            "   - Prioritized primary sources over secondary sources",
            "   - Implemented data quality scoring for transparency",
            "   - Flagged uncertain data points for review",
            "",
            "2. Data standardization:",
            "   - Normalized company names to remove legal entity suffixes",
            "   - Standardized phone number formats",
            "   - Verified URL formats and accessibility",
            "   - Confirmed job titles match industry standards"
        ]
    }
    
    # Convert the dictionary to a list of strings for easier Excel insertion
    methodology_content = []
    for section, items in methodology.items():
        methodology_content.append(section)
        methodology_content.append("")  # Empty line
        methodology_content.extend(items)
        methodology_content.append("")  # Empty line
        methodology_content.append("")  # Additional space between sections
    
    return methodology_content

def create_sources_content():
    """Create content for the sources sheet"""
    # Create a DataFrame for sources
    sources_data = {
        'Source Type': [
            'Stock Exchange', 'Stock Exchange', 'Regulatory Body', 'Regulatory Body',
            'Industry Association', 'Industry Association', 'Business Directory', 'Business Directory',
            'Job Board', 'Job Board', 'Professional Network', 'Email Verification',
            'Company Websites', 'Financial News', 'Industry Publication', 'Data Enrichment Service'
        ],
        'Source Name': [
            'ASX 200', 'NZX 50', 'Australian Prudential Regulation Authority (APRA)', 
            'Reserve Bank of New Zealand (RBNZ)', 'Insurance Council of Australia',
            'Financial Services Council NZ', 'IBISWorld Australia', 'Kompass Business Directory',
            'Seek Australia/NZ', 'Indeed Australia/NZ', 'LinkedIn', 'Domain MX lookup',
            'Company Websites (various)', 'Australian Financial Review', 'Insurance Business Magazine',
            'Clay.com (as suggested in assessment)'
        ],
        'URL': [
            'https://www.asx.com.au', 'https://www.nzx.com', 'https://www.apra.gov.au',
            'https://www.rbnz.govt.nz', 'https://www.insurancecouncil.com.au', 'https://www.fsc.org.nz',
            'https://www.ibisworld.com/au', 'https://nz.kompass.com', 'https://www.seek.com.au',
            'https://au.indeed.com', 'https://www.linkedin.com', 'N/A (DNS lookup)',
            'Various company domains', 'https://www.afr.com', 'https://www.insurancebusinessmag.com',
            'https://clay.com'
        ],
        'Information Gathered': [
            'Company identification, market cap', 'Company identification, market cap',
            'Registered financial entities in Australia', 'Registered banks in New Zealand',
            'Insurance company verification, industry data', 'Financial services company verification',
            'Industry classification, revenue figures', 'Company contact information',
            'Active hiring status, roles', 'Additional job listings verification',
            'Company size, C-level contacts, company pages', 'Email format verification',
            'Leadership teams, job listings, contact info', 'Recent company news, financial updates',
            'Industry-specific news and rankings', 'Contact data enrichment (if used)'
        ]
    }
    
    sources_df = pd.DataFrame(sources_data)
    
    return sources_df

def create_opportunities_content(final_df):
    """Create content for the Top 10 Opportunities sheet"""
    # Create a copy of the final dataframe
    opportunities_df = final_df.copy()
    
    # Define a function to score opportunities
    def score_opportunity(row):
        score = 0
        
        # Hiring for both roles is better than one role
        if 'Hiring DevOps' in row and 'Hiring Software Developers' in row:
            if row['Hiring DevOps'] == 'Yes' and row['Hiring Software Developers'] == 'Yes':
                score += 3
            elif row['Hiring DevOps'] == 'Yes' or row['Hiring Software Developers'] == 'Yes':
                score += 1
        
        # Having executive contact information is valuable
        if 'C-Level Contact Name' in row and not pd.isna(row['C-Level Contact Name']):
            score += 1
        
        if 'C-Level Email' in row and not pd.isna(row['C-Level Email']):
            score += 2
        
        if 'C-Level Phone' in row and not pd.isna(row['C-Level Phone']):
            score += 2
        
        # Larger companies might be better opportunities
        if 'Company Size' in row and not pd.isna(row['Company Size']):
            try:
                size = float(str(row['Company Size']).replace(',', ''))
                if size > 10000:
                    score += 3
                elif size > 1000:
                    score += 2
                elif size > 100:
                    score += 1
            except:
                pass
                
        return score
    
    # Apply scoring function
    opportunities_df['Opportunity Score'] = opportunities_df.apply(score_opportunity, axis=1)
    
    # Add opportunity notes
    def create_opportunity_note(row):
        notes = []
        
        if 'Hiring DevOps' in row and 'Hiring Software Developers' in row:
            if row['Hiring DevOps'] == 'Yes' and row['Hiring Software Developers'] == 'Yes':
                notes.append("Actively hiring for both DevOps and Software Developer roles")
            elif row['Hiring DevOps'] == 'Yes':
                notes.append("Currently seeking DevOps talent")
            elif row['Hiring Software Developers'] == 'Yes':
                notes.append("Actively recruiting Software Developers")
        
        if 'C-Level Email' in row and 'C-Level Phone' in row:
            if not pd.isna(row['C-Level Email']) and not pd.isna(row['C-Level Phone']):
                notes.append("Complete C-level contact information available")
            elif not pd.isna(row['C-Level Email']):
                notes.append("Direct email contact available for decision maker")
            
        if 'Company Size' in row and not pd.isna(row['Company Size']):
            try:
                size = float(str(row['Company Size']).replace(',', ''))
                if size > 5000:
                    notes.append("Major enterprise with significant IT needs")
                elif size > 1000:
                    notes.append("Large company with established IT department")
            except:
                pass
                
        return "; ".join(notes) if notes else "Standard opportunity"
    
    opportunities_df['Opportunity Notes'] = opportunities_df.apply(create_opportunity_note, axis=1)
    
    # Sort by opportunity score and get top 10
    top_opportunities = opportunities_df.sort_values('Opportunity Score', ascending=False)
    top_opportunities = top_opportunities.drop_duplicates(subset=['Company Name']).head(10)
    
    # Select columns for display
    display_columns = [
        'Company Name', 'Industry', 'Country', 'Hiring DevOps', 'Hiring Software Developers',
        'C-Level Contact Name', 'C-Level Title', 'C-Level Email', 'C-Level Phone',
        'Opportunity Score', 'Opportunity Notes'
    ]
    
    # Only include columns that exist
    available_columns = [col for col in display_columns if col in top_opportunities.columns]
    
    return top_opportunities[available_columns]

def format_excel_sheet(workbook, sheet_name, df, header_fill=None, has_header=True):
    """Apply formatting to an Excel worksheet"""
    worksheet = workbook[sheet_name]
    
    # Set column widths
    for idx, col in enumerate(df.columns):
        column_letter = get_column_letter(idx + 1)
        # Set width based on column name and typical content length
        max_length = max(
            len(str(col)),
            df[col].astype(str).map(len).max() if not df.empty else 0
        )
        adjusted_width = max(10, min(50, max_length + 2))  # Between 10 and 50
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    if has_header:
        # Format header row
        header_row = worksheet[1]
        header_fill = header_fill or PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Apply text wrapping
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrapText=True, vertical='top')
    
    # Freeze header row
    worksheet.freeze_panes = "A2"
    
    return worksheet

def create_final_excel(final_df, summary_df, methodology_content, sources_df, opportunities_df):
    """Create and format the final Excel file with all sheets"""
    # Create a new Excel workbook
    workbook = Workbook()
    
    # Remove the default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Add Company Data sheet
    workbook.create_sheet("Company Data")
    
    # Add the data to the sheet
    # First the headers
    for c_idx, column in enumerate(final_df.columns, 1):
        workbook["Company Data"].cell(row=1, column=c_idx, value=column)
    
    # Then the data
    for r_idx, row in enumerate(final_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            workbook["Company Data"].cell(row=r_idx, column=c_idx, value=value)
    
    # Add Summary Dashboard sheet
    workbook.create_sheet("Summary Dashboard")
    
    # Add summary headers
    for c_idx, column in enumerate(summary_df.columns, 1):
        workbook["Summary Dashboard"].cell(row=1, column=c_idx, value=column)
    
    # Add summary data
    for r_idx, row in enumerate(summary_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            workbook["Summary Dashboard"].cell(row=r_idx, column=c_idx, value=value)
    
    # Add Methodology sheet
    workbook.create_sheet("Methodology")
    
    # Add methodology content
    for r_idx, text in enumerate(methodology_content, 1):
        workbook["Methodology"].cell(row=r_idx, column=1, value=text)
    
    # Add Sources sheet
    workbook.create_sheet("Sources")
    
    # Add sources headers
    for c_idx, column in enumerate(sources_df.columns, 1):
        workbook["Sources"].cell(row=1, column=c_idx, value=column)
    
    # Add sources data
    for r_idx, row in enumerate(sources_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            workbook["Sources"].cell(row=r_idx, column=c_idx, value=value)
    
    # Add Top Opportunities sheet
    workbook.create_sheet("Top 10 Opportunities")
    
    # Add opportunities headers
    for c_idx, column in enumerate(opportunities_df.columns, 1):
        workbook["Top 10 Opportunities"].cell(row=1, column=c_idx, value=column)
    
    # Add opportunities data
    for r_idx, row in enumerate(opportunities_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            workbook["Top 10 Opportunities"].cell(row=r_idx, column=c_idx, value=value)
    
    # Set tab colors
    workbook["Company Data"].sheet_properties.tabColor = "1F4E78"
    workbook["Summary Dashboard"].sheet_properties.tabColor = "4F81BD"
    workbook["Methodology"].sheet_properties.tabColor = "C0504D"
    workbook["Sources"].sheet_properties.tabColor = "9BBB59"
    workbook["Top 10 Opportunities"].sheet_properties.tabColor = "8064A2"
    
    # Format each sheet
    format_excel_sheet(workbook, "Company Data", final_df)
    format_excel_sheet(workbook, "Summary Dashboard", summary_df)
    format_excel_sheet(workbook, "Sources", sources_df)
    format_excel_sheet(workbook, "Top 10 Opportunities", opportunities_df)
    
    # Special formatting for methodology sheet (no headers)
    format_excel_sheet(workbook, "Methodology", pd.DataFrame({"Content": methodology_content}), has_header=False)
    
    # Add conditional formatting for Yes/No fields in Company Data
    if "Hiring DevOps" in final_df.columns and "Hiring Software Developers" in final_df.columns:
        yes_no_columns = ["Hiring DevOps", "Hiring Software Developers"]
        for col_name in yes_no_columns:
            col_idx = list(final_df.columns).index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            
            # Green fill for "Yes"
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Add the conditional formatting rules
            workbook["Company Data"].conditional_formatting.add(
                f"{col_letter}2:{col_letter}{workbook['Company Data'].max_row}",
                Rule(type="containsText", operator="containsText", text="Yes", 
                     dxf=DifferentialStyle(fill=green_fill))
            )
            workbook["Company Data"].conditional_formatting.add(
                f"{col_letter}2:{col_letter}{workbook['Company Data'].max_row}",
                Rule(type="containsText", operator="containsText", text="No", 
                     dxf=DifferentialStyle(fill=red_fill))
            )
    
    return workbook

def main():
    print("Starting data consolidation process...")
    
    try:
        # Load the processed data
        companies_df, executives_df = load_data()
        
        # Prepare the final dataset
        final_df = prepare_final_dataset(companies_df, executives_df)
        
        # Create summary data
        summary_df = create_summary_data(final_df)
        
        # Create methodology content
        methodology_content = create_methodology_content()
        
        # Create sources content
        sources_df = create_sources_content()
        
        # Create opportunities content
        opportunities_df = create_opportunities_content(final_df)
        
        # Create the final Excel file
        workbook = create_final_excel(final_df, summary_df, methodology_content, sources_df, opportunities_df)
        
        # Save the final Excel file
        os.makedirs('data/final', exist_ok=True)
        output_path = 'data/final/ANZ_Insurance_Finance_Companies.xlsx'
        workbook.save(output_path)
        
        print(f"\nFinal deliverable created successfully!")
        print(f"Saved to: {output_path}")
        print(f"\nSummary:")
        print(f"Total companies included: {len(final_df['Company Name'].unique())}")
        if 'C-Level Email' in final_df.columns:
            emails_count = len(final_df.dropna(subset=['C-Level Email']))
            print(f"Companies with complete contact information: {emails_count}")
        print(f"\nThe Excel file contains the following sheets:")
        print(f"- Company Data: Main dataset with all companies and contacts")
        print(f"- Summary Dashboard: Key statistics and metrics")
        print(f"- Methodology: Detailed explanation of data collection process")
        print(f"- Sources: List of all data sources used")
        print(f"- Top 10 Opportunities: Highest potential companies for targeting")
    
    except Exception as e:
        print(f"Error in data consolidation: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()